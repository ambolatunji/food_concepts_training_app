import streamlit as st
import pandas as pd
import plotly.express as px
import io
import io, re
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from core.db import (
    get_conn, list_trainings, trainings_summary,
    distinct_employee_values, distinct_training_values,
    count_employees, count_employees_matching, count_trained_employees,
    count_joins, count_leaves, turnover_rate, employees_due_flags,
    training_frequency_report, stores_trained_count, stores_pending_training,
    staff_eligible_for_training, staff_participated_in_training,
    staff_yet_to_be_trained, trainings_pending_count, food_handlers_report,
    trained_employees_who_left,
    REQUIRED_TRAININGS
)
from core.report_generator import create_pptx_report, create_word_report, create_pdf_report


# Canonical buckets (must match your title bucketing)
TITLE_BUCKETS = {
    "food safety": "Food Safety Training",
    "fire": "Fire Safety Training",
    "first aid": "First Aid Training",
    "pest": "Pest Control Training",
    "occupational health": "Occupational Health and Safety Training",
    "ohs": "Occupational Health and Safety Training",
    "6s": "6S Training",
    "water treatment": "Water Treatment Plant Training",
}

# Header colors (hex without #). Tweak to your taste.
COLOR_MAP = {
    "Food Safety Training": "B7DEE8",   # light teal
    "Fire Safety Training": "C6E2FF",   # light blue
    "First Aid Training": "F7C6C7",     # light pink
    "Pest Control Training": "C6E0B4",  # light green
    "Occupational Health and Safety Training": "E2EFDA",
    "6S Training": "FFF2CC",            # light yellow
    "Water Treatment Plant Training": "D9D2E9",  # light violet
}

def bucket_title(title: str) -> str:
    if not title: return ""
    t = str(title).lower()
    for key, canon in TITLE_BUCKETS.items():
        if key in t:
            return canon
    return title

def _matrix_base(conn) -> pd.DataFrame:
    """Returns your wide matrix (flat columns) you already built, but robust if empty."""
    # employees
    emps = pd.read_sql_query("""
        SELECT id, employee_code AS "EMPLOYEE CODE",
               name AS "EMPLOYEE NAME",
               store AS "LOCATION",
               department AS "DIVISION"
        FROM employees
        WHERE deleted_at IS NULL
        ORDER BY name COLLATE NOCASE
    """, conn)

    # trainings
    trs = pd.read_sql_query("""
        SELECT t.employee_id, t.training_date, t.next_due_date, t.training_title
        FROM trainings t
        JOIN employees e ON e.id=t.employee_id
        WHERE t.deleted_at IS NULL AND e.deleted_at IS NULL
    """, conn)
    hires = pd.read_sql_query("""
        SELECT employee_id, MAX(event_date) AS hire_date
        FROM employment_events
        WHERE event_type='hire'
        GROUP BY employee_id
    """, conn)

    if trs.empty:
        out = emps.copy()
        out.insert(0, "S/N", range(1, len(out)+1))
        # add empty blocks so headers exist
        for canon in COLOR_MAP.keys():
            for sub in ("DATE TRAINED","DUE DATE","NO OF DAYS PRIOR DUE DATE","STATUS"):
                out[f"{canon} - {sub}"] = ""
        return out

    trs["bucket"] = trs["training_title"].apply(bucket_title)
    trs["training_date"] = pd.to_datetime(trs["training_date"])
    trs.sort_values(["employee_id","bucket","training_date"], inplace=True)
    last = trs.groupby(["employee_id","bucket"], as_index=False).last()

    today = pd.Timestamp(date.today())
    last["next_due_date"] = pd.to_datetime(last["next_due_date"])
    last["days_prior_due"] = (last["next_due_date"].dt.normalize() - today.normalize()).dt.days
    last["status"] = last["days_prior_due"].apply(lambda d: "NOT DUE" if pd.notna(d) and d > 0 else "DUE")

    # assemble
    out = emps.copy()
    for canon in COLOR_MAP.keys():
        sub = last[last["bucket"] == canon][["employee_id","training_date","next_due_date","days_prior_due","status"]].copy()
        sub.columns = ["employee_id",
                       f"{canon} - DATE TRAINED",
                       f"{canon} - DUE DATE",
                       f"{canon} - NO OF DAYS PRIOR DUE DATE",
                       f"{canon} - STATUS"]
        out = out.merge(sub, left_on="id", right_on="employee_id", how="left")
        if "employee_id" in out.columns:
            out.drop(columns=["employee_id"], inplace=True)

    out.insert(0, "S/N", range(1, len(out)+1))
    out = out.merge(hires, left_on="id", right_on="employee_id", how="left")  # hire_date per employee (NaN if not a new hire)
    today = pd.Timestamp(pd.Timestamp.today().date())

    # probation window = 183 days
    hire_dt = pd.to_datetime(out["hire_date"], errors="coerce")
    on_probation = (today - hire_dt.dt.normalize()).dt.days < 183
    # Important: ONLY rows with a hire_date count as "new hire". If no hire_date → NOT on probation.
    on_probation = on_probation & hire_dt.notna()

    for canon in COLOR_MAP.keys():
        dcol = f"{canon} - DATE TRAINED"
        ncol = f"{canon} - DUE DATE"
        pcol = f"{canon} - NO OF DAYS PRIOR DUE DATE"
        scol = f"{canon} - STATUS"
        if dcol in out.columns:
            # Missing training date?
            mask_missing = out[dcol].isna()
            # We fill 1900-01-01 for everyone missing EXCEPT employees currently on probation
            apply_mask = mask_missing & (~on_probation)
            out.loc[apply_mask, dcol] = pd.Timestamp("1900-01-01")
            out.loc[apply_mask, ncol] = today            # due now
            out.loc[apply_mask, pcol] = 0
            out.loc[apply_mask, scol] = "DUE"

    # cleanup helper columns
    for c in ("employee_id", "hire_date"):
        if c in out.columns:
            out.drop(columns=[c], inplace=True, errors="ignore")
    return out

def build_matrix_multiindex(conn) -> pd.DataFrame:
    """Convert flat matrix columns into a 2-row MultiIndex suitable for merged headers in Excel."""
    df = _matrix_base(conn)
    tuples = []
    for c in df.columns:
        if " - " in c:
            top, sub = c.split(" - ", 1)
            tuples.append((top, sub))
        else:
            tuples.append(("", c))
    df.columns = pd.MultiIndex.from_tuples(tuples)
    return df

def style_matrix_preview(df_mi: pd.DataFrame):
    """Preview in Streamlit with light background tint per training block (no merged headers in-browser)."""
    # flatten header with a newline so both levels are visible
    df_disp = df_mi.copy()
    df_disp.columns = [ (f"{a}\n{s}" if a else s) for a,s in df_mi.columns.to_list() ]
    stl = df_disp.style
    # tint blocks
    for train, hexcol in COLOR_MAP.items():
        # all columns where top level == train
        cols = [df_disp.columns[i] for i, (a, s) in enumerate(df_mi.columns) if a == train]
        if cols:
            stl = stl.set_properties(subset=cols, **{"background-color": f"#{hexcol}"})
    # header styling
    stl = stl.set_table_styles([{
        "selector": "th",
        "props": [("background-color","#2A2A2A"), ("color","white"), ("font-weight","bold")]
    }])
    return stl

def export_matrix_excel(df_mi, sheet_name="Matrix") -> bytes:
    """
    Write a MultiIndex-column matrix to Excel with merged, colored training headers.
    Works even when pandas can't write MultiIndex headers with index=False.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # --- styles ---
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df_mi.columns)  # list of (top, sub)
    ncols = len(cols)

    # 1) Build headers (row 1 = training group, row 2 = subheaders)
    j = 1  # Excel column index
    i = 0
    while i < ncols:
        top, sub = cols[i]
        if top == "":  # static column, merge vertically rows 1..2
            ws.merge_cells(start_row=1, start_column=j, end_row=2, end_column=j)
            c = ws.cell(row=1, column=j)
            c.value = sub
            c.font = bold
            c.alignment = center if sub != "EMPLOYEE NAME" else left
            c.border = border
            j += 1
            i += 1
            continue

        # training block: find contiguous columns with same top
        k = i
        while k < ncols and cols[k][0] == top:
            k += 1
        start_col = j
        end_col = j + (k - i) - 1

        # merge header row 1 across the block
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        c = ws.cell(row=1, column=start_col)
        c.value = top
        c.font = bold
        c.alignment = center

        # color header rows for the whole block
        hexcol = COLOR_MAP.get(top, "DDDDDD")
        fill = PatternFill(start_color=hexcol, end_color=hexcol, fill_type="solid")

        # write row-2 subheaders for the block
        for col_off, idx in enumerate(range(i, k)):
            sublabel = cols[idx][1]
            h = ws.cell(row=2, column=start_col + col_off)
            h.value = sublabel
            h.font = bold
            h.alignment = center
            h.fill = fill
            # borders top/bottom on both header rows
            ws.cell(row=1, column=start_col + col_off).fill = fill
            ws.cell(row=1, column=start_col + col_off).border = border
            ws.cell(row=2, column=start_col + col_off).border = border

        j = end_col + 1
        i = k

    # 2) Write data starting row 3 (no index)
    def _write_cell(r, c, val, sub_label=""):
        cell = ws.cell(row=r, column=c)
        if val is None or (isinstance(val, float) and pd.isna(val)):
            cell.value = None
        else:
            # date-like formatting
            if isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime().date()
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(val, (datetime, date)):
                cell.value = val
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(val, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", val):
                # parse simple ISO date string to date
                try:
                    y, m, d = map(int, val.split("-"))
                    cell.value = date(y, m, d)
                    cell.number_format = "yyyy-mm-dd"
                except Exception:
                    cell.value = val
            else:
                cell.value = val
        # align text columns
        # IMPORTANT: always assign a *new* Alignment object (never reuse cell.alignment)
        if sub_label in ("EMPLOYEE NAME","LOCATION","DIVISION"):
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = border

    start_row = 3
    for r_idx, row in enumerate(df_mi.itertuples(index=False, name=None), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            top, sub = cols[c_idx - 1]
            _write_cell(r_idx, c_idx, val, sub_label=sub if top == "" else "")

    # 3) For top-level empty columns, write their header text (we already put it row 1)
    #    For training blocks, row-2 subheaders already set above.

    # 4) Freeze header
    ws.freeze_panes = ws["A3"]

    # 5) Column widths
    width_map = {
        "S/N": 6, "EMPLOYEE CODE": 14, "EMPLOYEE NAME": 24, "LOCATION": 18, "DIVISION": 18,
        "DATE TRAINED": 16, "DUE DATE": 16, "NO OF DAYS PRIOR DUE DATE": 24, "STATUS": 12,
    }
    for col_idx, (top, sub) in enumerate(cols, start=1):
        hdr = sub if top == "" else sub
        w = width_map.get(hdr, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 6) Return bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def to_excel_bytes(df_dict: dict, filename="export.xlsx"):
    # df_dict = {"SheetName": dataframe, ...}
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        for sheet, df in df_dict.items():
            # openpyxl sheet names max 31 chars
            safe = (sheet or "Sheet")[:31]
            df.to_excel(writer, index=False, sheet_name=safe)
    bio.seek(0)
    return bio.getvalue()
def to_excel_bytes(sheets: dict) -> bytes:
    """sheets={'SheetName': df, ...} → Excel bytes"""
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        for name, df in sheets.items():
            df.to_excel(xw, index=False, sheet_name=(name or "Sheet")[:31])
    bio.seek(0)
    return bio.getvalue()

def get_trainings_df(conn, filters:dict) -> pd.DataFrame:
    rows = list_trainings(conn, filters)
    df = pd.DataFrame(rows, columns=[
        "Training ID","Employee Name","Department","Store","Position","Region",
        "Training Date","Next Due Date","Evidence Path","Training Title","Training Venue"
    ])
    if not df.empty:
        df["Training Date"] = pd.to_datetime(df["Training Date"], errors="coerce")
        df["Next Due Date"] = pd.to_datetime(df["Next Due Date"], errors="coerce")
    return df

def get_employees_df(conn, filters:dict) -> pd.DataFrame:
    sql = """SELECT id, name, department, store, position, region, start_date, end_date
             FROM employees WHERE deleted_at IS NULL"""
    args=[]
    for k in ("department","store","position","region"):
        if filters.get(k):
            sql += f" AND lower({k})=?"; args.append(filters[k].lower())
    return pd.read_sql_query(sql, conn, params=args)

def chart_trainings_by_department(df):
    if df.empty: return None, pd.DataFrame()
    g = df.groupby("Department")["Training ID"].count().reset_index(name="Trainings").sort_values("Trainings", ascending=False)
    fig = px.bar(g, x="Department", y="Trainings", title="Trainings by Department")
    fig.update_layout(margin=dict(l=30,r=10,b=120,t=60), xaxis_tickangle=-45)
    return fig, g

def chart_trainings_by_title(df):
    if df.empty: return None, pd.DataFrame()
    g = df.groupby("Training Title")["Training ID"].count().reset_index(name="Trainings").sort_values("Trainings", ascending=False)
    fig = px.bar(g, x="Training Title", y="Trainings", title="Trainings by Title")
    fig.update_layout(margin=dict(l=30,r=10,b=160,t=60), xaxis_tickangle=-45)
    return fig, g

def chart_trainings_by_month(df):
    if df.empty: return None, pd.DataFrame()
    d = df.dropna(subset=["Training Date"]).copy()
    if d.empty: return None, pd.DataFrame()
    d["Month"] = d["Training Date"].dt.to_period("M").astype(str)
    g = d.groupby("Month")["Training ID"].count().reset_index(name="Trainings")
    fig = px.line(g, x="Month", y="Trainings", title="Training Trend by Month", markers=True)
    fig.update_layout(margin=dict(l=30,r=10,b=80,t=60))
    return fig, g

def chart_region_pie(df):
    if df.empty: return None, pd.DataFrame()
    g = df.groupby("Region")["Training ID"].count().reset_index(name="Trainings")
    if g.empty: return None, g
    fig = px.pie(g, names="Region", values="Trainings", title="Trainings Share by Region", hole=0.35)
    fig.update_layout(margin=dict(l=10,r=10,b=10,t=60))
    return fig, g

def chart_trainings_by_region(df):
    if df.empty: return None, pd.DataFrame()
    g = df.groupby("Region")["Training ID"].count().reset_index(name="Trainings").sort_values("Trainings", ascending=False)
    fig = px.bar(g, x="Region", y="Trainings", title="Trainings by Region", text="Trainings")
    fig.update_layout(margin=dict(l=30,r=10,b=120,t=60), xaxis_tickangle=-45)
    return fig, g

def chart_top_stores(df, topn=20):
    if df.empty: return None, pd.DataFrame()
    g = df.groupby("Store")["Training ID"].count().reset_index(name="Trainings").sort_values("Trainings", ascending=False).head(topn)
    fig = px.bar(g, x="Store", y="Trainings", title=f"Top {topn} Stores by Trainings")
    fig.update_layout(margin=dict(l=30,r=10,b=200,t=60), xaxis_tickangle=-65)
    return fig, g

def chart_heat_dept_month(df):
    if df.empty or "Training Date" not in df.columns: return None, pd.DataFrame()
    d = df.dropna(subset=["Training Date"]).copy()
    if d.empty: return None, pd.DataFrame()
    d["Month"] = d["Training Date"].dt.to_period("M").astype(str)
    pvt = d.pivot_table(index="Department", columns="Month", values="Training ID", aggfunc="count", fill_value=0)
    if pvt.empty: return None, pvt
    fig = px.imshow(pvt, aspect="auto", title="Heatmap: Trainings by Department × Month", color_continuous_scale="Blues")
    fig.update_layout(margin=dict(l=40,r=20,b=120,t=60), xaxis_tickangle=-45)
    return fig, pvt.reset_index()

def chart_stacked_titles_by_dept(df, top_titles=6):
    if df.empty: return None, pd.DataFrame()
    top = df["Training Title"].value_counts().head(top_titles).index.tolist()
    d = df[df["Training Title"].isin(top)]
    if d.empty: return None, pd.DataFrame()
    g = d.groupby(["Department","Training Title"])["Training ID"].count().reset_index(name="Trainings")
    fig = px.bar(g, x="Department", y="Trainings", color="Training Title", barmode="stack",
                 title=f"Stacked Trainings by Department (Top {top_titles} Titles)")
    fig.update_layout(margin=dict(l=30,r=10,b=160,t=60), xaxis_tickangle=-45)
    return fig, g

def chart_coverage_by_department(conn, df, filters):
    emps = get_employees_df(conn, filters)
    if emps.empty:
        cov = pd.DataFrame(columns=["Department","Headcount","Trained","Coverage"])
        return None, cov
    head = emps.groupby("department")["id"].nunique().reset_index(name="Headcount")
    if df.empty:
        cov = head.assign(Trained=0, Coverage=0.0)
    else:
        trained = df.groupby("Department")["Employee Name"].nunique().reset_index(name="Trained")
        cov = head.merge(trained, left_on="department", right_on="Department", how="left").fillna({"Trained":0})
        cov["Coverage"] = (cov["Trained"]/cov["Headcount"]*100.0).round(1)
        # Drop the redundant Department column from merge, keep department
        if "Department" in cov.columns and "department" in cov.columns:
            cov = cov.drop(columns=["Department"])
    cov.rename(columns={"department":"Department"}, inplace=True)
    if cov.empty: return None, cov
    fig = px.bar(cov.sort_values("Coverage", ascending=False), x="Department", y="Coverage", title="Coverage by Department (%)")
    fig.update_layout(margin=dict(l=30,r=10,b=160,t=60), xaxis_tickangle=-45, yaxis_range=[0,100])
    return fig, cov

def chart_due_overdue_by_dept(conn, filters):
    emps = get_employees_df(conn, filters)[["id","department"]].rename(columns={"department":"Department"})
    if emps.empty: return None, pd.DataFrame()
    fl = employees_due_flags(conn, filters)
    if not fl:
        g = emps.assign(**{"Due≤30d":0, "Overdue":0}).groupby("Department")[["Due≤30d","Overdue"]].sum().reset_index()
    else:
        fdf = pd.DataFrame([{"id":k, **v} for k,v in fl.items()])
        base = emps.merge(fdf, on="id", how="left").fillna({"any_due_30":0,"any_overdue":0})
        g = base.groupby("Department")[["any_due_30","any_overdue"]].sum().reset_index().rename(
            columns={"any_due_30":"Due≤30d","any_overdue":"Overdue"})
    fig = px.bar(g.sort_values("Due≤30d", ascending=False), x="Department", y=["Due≤30d","Overdue"],
                 barmode="stack", title="Due vs Overdue by Department")
    fig.update_layout(margin=dict(l=30,r=10,b=160,t=60), xaxis_tickangle=-45)
    return fig, g

def chart_treemap_store_coverage(conn, df, filters):
    emps = get_employees_df(conn, filters)[["id","region","store"]]
    if emps.empty: return None, pd.DataFrame()
    head = emps.groupby(["region","store"])["id"].nunique().reset_index(name="Headcount")
    if df.empty:
        cov = head.assign(Trained=0, Coverage=0.0)
    else:
        trained = df.groupby("Store")["Employee Name"].nunique().reset_index(name="Trained")
        cov = head.merge(trained, left_on="store", right_on="Store", how="left").fillna({"Trained":0})
        cov["Coverage"] = (cov["Trained"]/cov["Headcount"]*100.0).round(1)
        # Drop redundant Store column from merge
        if "Store" in cov.columns and "store" in cov.columns:
            cov = cov.drop(columns=["Store"])
    cov.rename(columns={"region":"Region","store":"Store"}, inplace=True)
    if cov.empty: return None, cov
    fig = px.treemap(cov, path=["Region","Store"], values="Headcount",
                     color="Coverage", color_continuous_scale="RdYlGn",
                     title="Store Coverage Treemap (size = headcount, color = coverage%)")
    fig.update_layout(margin=dict(l=10,r=10,b=10,t=60))
    return fig, cov

def chart_joins_vs_leaves_by_month(conn, filters, date_from:str=None, date_to:str=None):
    emp = get_employees_df(conn, filters)[["start_date","end_date"]].copy()
    if emp.empty: return None, pd.DataFrame()
    emp["start_date"] = pd.to_datetime(emp["start_date"], errors="coerce")
    emp["end_date"]   = pd.to_datetime(emp["end_date"], errors="coerce")
    if date_from: emp = emp[(emp["start_date"].isna()) | (emp["start_date"]>=pd.to_datetime(date_from))]
    if date_to:   emp = emp[(emp["start_date"].isna()) | (emp["start_date"]<=pd.to_datetime(date_to))]
    j = emp.dropna(subset=["start_date"]).assign(Month=lambda d: d["start_date"].dt.to_period("M").astype(str)).groupby("Month").size().reset_index(name="Joins")
    l = emp.dropna(subset=["end_date"]).assign(Month=lambda d: d["end_date"].dt.to_period("M").astype(str)).groupby("Month").size().reset_index(name="Leaves")
    g = pd.merge(j, l, on="Month", how="outer").fillna(0).sort_values("Month")
    fig = px.bar(g, x="Month", y=["Joins","Leaves"], barmode="group", title="Joins vs Leaves by Month")
    fig.update_layout(margin=dict(l=30,r=10,b=80,t=60))
    return fig, g

conn = get_conn()
st.title("📊 Dashboard")
st.caption("Filter by dimensions and date range to analyze training coverage and trends.")

regions = ["All"] + distinct_employee_values(conn, "region")
positions = ["All"] + distinct_employee_values(conn, "position")
stores = ["All"] + distinct_employee_values(conn, "store")
departments = ["All"] + distinct_employee_values(conn, "department")
titles = ["All"] + distinct_training_values(conn, "training_title")
venues = ["All"] + distinct_training_values(conn, "training_venue")

c1, c2, c3, c4 = st.columns(4)
with c1:
    region = st.selectbox("Region", regions)
with c2:
    department = st.selectbox("Department", departments)
with c3:
    store = st.selectbox("Store", stores)
with c4:
    position = st.selectbox("Position", positions)

c5, c6 = st.columns(2)
with c5:
    f_title = st.selectbox("Training Title", titles)
with c6:
    f_venue = st.selectbox("Training Venue", venues)

c7, c8 = st.columns(2)
with c7:
    date_from = st.date_input("From", value=None, format="YYYY-MM-DD")
with c8:
    date_to = st.date_input("To", value=None, format="YYYY-MM-DD")

filters={}
if region!="All": filters["region"]=region
if department!="All": filters["department"]=department
if store!="All": filters["store"]=store
if position!="All": filters["position"]=position
if date_from: filters["date_from"]=date_from.isoformat()
if date_to: filters["date_to"]=date_to.isoformat()
if f_title!="All": filters["training_title"]=f_title
if f_venue!="All": filters["training_venue"]=f_venue

rows = list_trainings(conn, filters=filters)
df = pd.DataFrame(rows, columns=[
    "Training ID","Employee Name","Department","Store","Position","Region",
    "Training Date","Next Due Date","Evidence Path","Training Title","Training Venue"
])

# --- KPIs (fully filter-aware) ---
# Total = Master + Hires + Leavers
# Active = Master + Hires - Leavers
total_emps_all = count_employees(conn, active_only=False)
active_emps_all = count_employees(conn, active_only=True)
total_emps_filtered = count_employees_matching(conn, filters)
trained_emps = count_trained_employees(conn, filters, active_only=True)  # Count only active employees trained
total_trainings = len(df)

# Due flags (probation-aware + "at least one training due" semantics)
flags = employees_due_flags(conn, filters)
any_due_30 = sum(1 for v in flags.values() if v["any_due_30"])
any_overdue = sum(1 for v in flags.values() if v["any_overdue"])
# Coverage = Trained Active Employees / Total Active Employees
coverage = (trained_emps / active_emps_all * 100.0) if active_emps_all else 0.0

# Joins/Leaves counters for selected date range (if none, default to “this year”)
from datetime import date
dfrom = filters.get("date_from") or f"{date.today().year}-01-01"
dto   = filters.get("date_to") or date.today().isoformat()
joins  = count_joins(conn, dfrom, dto, filters)
leaves = count_leaves(conn, dfrom, dto, filters)
tvr    = turnover_rate(conn, dfrom, dto, filters)

st.markdown("""
<style>
.kpi {border-radius:16px; padding:16px; background:rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12);}
.kpi h3 {margin:0 0 6px 0; font-size:0.9rem; opacity:0.8;}
.kpi p {margin:0; font-size:1.6rem; font-weight:700;}
.kpi small {opacity:0.7;}
</style>
""", unsafe_allow_html=True)

kc1, kc2, kc3, kc4, kc5, kc6 = st.columns(6)
with kc1: st.markdown(f'<div class="kpi"><h3>Total Trainings</h3><p>{total_trainings}</p></div>', unsafe_allow_html=True)
with kc2: st.markdown(f'<div class="kpi"><h3>Active Employees</h3><p>{active_emps_all}</p><small>Total: {total_emps_all}</small></div>', unsafe_allow_html=True)
with kc3: st.markdown(f'<div class="kpi"><h3>Employees Trained</h3><p>{trained_emps}</p><small>Active only</small></div>', unsafe_allow_html=True)
with kc4: st.markdown(f'<div class="kpi"><h3>Coverage</h3><p>{coverage:.1f}%</p><small>Trained/Active</small></div>', unsafe_allow_html=True)
with kc5: st.markdown(f'<div class="kpi"><h3>Due ≤ 30d</h3><p>{any_due_30}</p><small>Overdue: {any_overdue}</small></div>', unsafe_allow_html=True)
with kc6: st.markdown(f'<div class="kpi"><h3>Joins / Leaves</h3><p>{joins} / {leaves}</p><small>Turnover: {tvr:.1f}%</small></div>', unsafe_allow_html=True)

#fig2 = px.bar(df, x=by.capitalize(), y="Trainings", title=f"Trainings by {by.capitalize()} (filtered)")

# Client-side filter for title/venue (simpler than altering SQL)
if f_title!="All":
    df = df[df["Training Title"]==f_title]
if f_venue!="All":
    df = df[df["Training Venue"]==f_venue]

st.subheader("Breakdown")
by = st.selectbox("Group by", ["department","store","region","position","training_title","training_venue"], index=0)
summary = trainings_summary(conn, filters.get("date_from"), filters.get("date_to"), by=by)
df_sum = pd.DataFrame(summary, columns=[by.capitalize(),"Trainings"])
if len(df_sum)>0:
    fig2 = px.bar(df_sum, x=by.capitalize(), y="Trainings", title=f"Trainings by {by.capitalize()}")
    title = f"Trainings by {by.capitalize()} (filtered)"
    st.plotly_chart(fig2, use_container_width=True)
st.dataframe(df, use_container_width=True)

st.subheader("Turnover analysis")
grp = st.selectbox("Group turnover by", ["Month","Department","Store","Region"], index=0)
import pandas as pd

def _month(s): 
    return pd.to_datetime(s).dt.to_period("M").astype(str)

# Build mini tables
joins_df = pd.read_sql_query("""
    SELECT name, department, store, region, start_date AS date
    FROM employees
    WHERE deleted_at IS NULL AND start_date IS NOT NULL
""", conn)
leaves_df = pd.read_sql_query("""
    SELECT name, department, store, region, end_date AS date
    FROM employees
    WHERE deleted_at IS NULL AND end_date IS NOT NULL
""", conn)

# Apply dimension filters
for k in ("department","store","region"):
    if filters.get(k):
        if not joins_df.empty:  joins_df = joins_df[joins_df[k].str.lower()==filters[k].lower()]
        if not leaves_df.empty: leaves_df = leaves_df[leaves_df[k].str.lower()==filters[k].lower()]

if not joins_df.empty:  joins_df["date"]  = pd.to_datetime(joins_df["date"])
if not leaves_df.empty: leaves_df["date"] = pd.to_datetime(leaves_df["date"])

if grp=="Month":
    if not joins_df.empty:  joins_tab = joins_df.assign(Month=_month(joins_df["date"])).groupby("Month").size().reset_index(name="Joins")
    else: joins_tab = pd.DataFrame(columns=["Month","Joins"])
    if not leaves_df.empty: leaves_tab = leaves_df.assign(Month=_month(leaves_df["date"])).groupby("Month").size().reset_index(name="Leaves")
    else: leaves_tab = pd.DataFrame(columns=["Month","Leaves"])
    turn_tab = pd.merge(joins_tab, leaves_tab, on="Month", how="outer").fillna(0).sort_values("Month")
    st.dataframe(turn_tab, use_container_width=True)
else:
    key = grp.lower()
    if not joins_df.empty:  joins_tab = joins_df.groupby(key).size().reset_index(name="Joins")
    else: joins_tab = pd.DataFrame(columns=[key,"Joins"])
    if not leaves_df.empty: leaves_tab = leaves_df.groupby(key).size().reset_index(name="Leaves")
    else: leaves_tab = pd.DataFrame(columns=[key,"Leaves"])
    turn_tab = pd.merge(joins_tab, leaves_tab, on=key, how="outer").fillna(0).sort_values(key)
    st.dataframe(turn_tab, use_container_width=True)

st.download_button(
    "Download table (CSV)",
    data=df.to_csv(index=False).encode("utf-8"),
    file_name="trainings_filtered.csv",
    mime="text/csv",
    use_container_width=True
)
st.download_button(
    "Download table (Excel)",
    data=to_excel_bytes({"Trainings": df}),
    file_name="trainings_filtered.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)
# Per-group multi-sheet workbook (each group's rows)
if not df.empty:
    group_frames = {}
    for g in df_sum[by.capitalize()].dropna().unique():
        group_frames[str(g)] = df[df[by.capitalize()] == g]
    if group_frames:
        st.download_button(
            f"Download per-{by} detail (multi-sheet Excel)",
            data=to_excel_bytes(group_frames),
            file_name=f"trainings_by_{by}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
st.caption(f"Total Employees (all time): {total_emps_all} | Trainings shown: {len(df)}")
from datetime import date

# canonical buckets for titles
st.subheader("Matrix")

# Build MultiIndex matrix once
matrix_mi = build_matrix_multiindex(conn)

# Preview controls
opt = st.radio("Rows to show", ["10", "20", "50","100","All"], horizontal=True, index=0)
n = None if opt=="All" else int(opt)

# Preview (styled)
preview = matrix_mi if n is None else matrix_mi.head(n)
#st.dataframe(style_matrix_preview(preview), use_container_width=True, height=520)
st.write(style_matrix_preview(preview))#.to_html(), unsafe_allow_html=True)
# Downloads
st.download_button(
    "Download styled Matrix (Excel)",
    data=export_matrix_excel(matrix_mi),
    file_name="Training Documents.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)
st.caption(f"Matrix generated on {date.today().isoformat()}")

# ---- NEW COMPREHENSIVE REPORTS SECTION ----
st.header("📋 Comprehensive Reports")
st.caption("Detailed analytics and insights for training management")

# Year selector for reports
report_year = st.selectbox("Select Year for Reports",
                           options=list(range(date.today().year, date.today().year - 5, -1)),
                           index=0)

# Download Full Comprehensive Report Section
st.markdown("### 📥 Download Full Comprehensive Report")
st.caption("Download all reports below in PowerPoint, Word, or PDF format")

col_dl1, col_dl2, col_dl3 = st.columns(3)

# Prepare report data for all formats
def prepare_report_data():
    freq_data = training_frequency_report(conn, year=report_year, filters=filters)
    freq_df = pd.DataFrame(freq_data, columns=["Training Title", "Times Held", "All Dates", "First Date", "Last Date"]) if freq_data else pd.DataFrame()

    stores_trained = stores_trained_count(conn, year=report_year, filters=filters)
    stores_pending = stores_pending_training(conn, year=report_year, filters=filters)
    total_stores_count = stores_trained + stores_pending

    eligible = staff_eligible_for_training(conn, filters)
    participated = staff_participated_in_training(conn, year=report_year, filters=filters)
    not_participated = eligible - participated

    food_report = food_handlers_report(conn, filters)
    pending_count = trainings_pending_count(conn, filters)

    # Fix: Get total trainings for the year, not from filtered df
    total_trainings = conn.execute("""
        SELECT COUNT(*) FROM trainings t
        JOIN employees e ON e.id = t.employee_id
        WHERE t.deleted_at IS NULL
          AND e.deleted_at IS NULL
          AND strftime('%Y', t.training_date) = ?
    """, (str(report_year),)).fetchone()[0]

    # Staff counts: Active vs Total
    total_staff = count_employees(conn, active_only=False)  # Total = Master + Hires + Leavers
    active_staff = count_employees(conn, active_only=True)  # Active = Master + Hires - Leavers
    staff_trained = count_trained_employees(conn, filters, active_only=True)  # Only active trained

    # Get compliance data
    compliance_data = []
    for training_title in REQUIRED_TRAININGS:
        training_count = conn.execute("""
            SELECT COUNT(DISTINCT t.employee_id)
            FROM trainings t
            JOIN employees e ON e.id = t.employee_id
            WHERE t.deleted_at IS NULL
              AND e.deleted_at IS NULL
              AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
              AND lower(t.training_title) LIKE ?
        """, (f'%{training_title.split()[0].lower()}%',)).fetchone()[0]

        compliance_rate = (training_count / eligible * 100) if eligible else 0

        overdue = conn.execute("""
            SELECT COUNT(DISTINCT t.employee_id)
            FROM trainings t
            JOIN employees e ON e.id = t.employee_id
            WHERE t.deleted_at IS NULL
              AND e.deleted_at IS NULL
              AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
              AND lower(t.training_title) LIKE ?
              AND date(t.next_due_date) < date('now')
        """, (f'%{training_title.split()[0].lower()}%',)).fetchone()[0]

        # Only add to compliance data if trained staff count > 0
        if training_count > 0 or compliance_rate > 0:
            compliance_data.append({
                'Training': training_title,
                'Trained Staff': training_count,
                'Compliance Rate': compliance_rate,
                'Overdue': overdue
            })

    # Get venue data
    venue_data = conn.execute("""
        SELECT t.training_venue,
               COUNT(*) as sessions,
               COUNT(DISTINCT t.employee_id) as unique_staff
        FROM trainings t
        JOIN employees e ON e.id = t.employee_id
        WHERE t.deleted_at IS NULL
          AND e.deleted_at IS NULL
          AND strftime('%Y', t.training_date) = ?
          AND t.training_venue IS NOT NULL
          AND t.training_venue != ''
        GROUP BY t.training_venue
        HAVING sessions > 0 AND unique_staff > 0
        ORDER BY sessions DESC
    """, (str(report_year),)).fetchall()
    venue_df = pd.DataFrame(venue_data, columns=['Venue', 'Sessions', 'Unique Staff']) if venue_data else pd.DataFrame()
    # Filter out rows with zero or negative values
    if not venue_df.empty:
        venue_df = venue_df[(venue_df['Sessions'] > 0) & (venue_df['Unique Staff'] > 0)]

    # Get regional data - using region from trainings table and filtering by year
    region_data = conn.execute("""
        SELECT COALESCE(t.region, e.region, 'Unknown') as region_name,
               COUNT(DISTINCT e.id) as total_staff,
               COUNT(DISTINCT CASE WHEN t.id IS NOT NULL THEN e.id END) as trained_staff,
               COUNT(CASE WHEN strftime('%Y', t.training_date) = ? THEN t.id END) as total_trainings
        FROM employees e
        LEFT JOIN trainings t ON e.id = t.employee_id
                              AND t.deleted_at IS NULL
                              AND strftime('%Y', t.training_date) = ?
        WHERE e.deleted_at IS NULL
          AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
        GROUP BY region_name
        HAVING total_trainings > 0
        ORDER BY total_trainings DESC
    """, (str(report_year), str(report_year))).fetchall()
    region_df = pd.DataFrame(region_data, columns=['Region', 'Total Staff', 'Trained Staff', 'Total Trainings']) if region_data else pd.DataFrame()
    # Filter out rows with zero or negative values
    if not region_df.empty:
        region_df = region_df[(region_df['Total Staff'] > 0) & (region_df['Total Trainings'] > 0)]

    # Get monthly data
    monthly_data = conn.execute("""
        SELECT strftime('%m', t.training_date) as month,
               COUNT(*) as trainings,
               COUNT(DISTINCT t.employee_id) as unique_staff
        FROM trainings t
        JOIN employees e ON e.id = t.employee_id
        WHERE t.deleted_at IS NULL
          AND e.deleted_at IS NULL
          AND strftime('%Y', t.training_date) = ?
        GROUP BY month
        HAVING trainings > 0
        ORDER BY month
    """, (str(report_year),)).fetchall()
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_df = pd.DataFrame(monthly_data, columns=['Month', 'Trainings', 'Unique Staff']) if monthly_data else pd.DataFrame()
    if not monthly_df.empty:
        # Filter out rows with zero or negative values
        monthly_df = monthly_df[(monthly_df['Trainings'] > 0) & (monthly_df['Unique Staff'] > 0)]
        monthly_df['Month Name'] = monthly_df['Month'].apply(lambda x: month_names[int(x)-1])

    # Get pending staff details
    yet_to_train = staff_yet_to_be_trained(conn, filters)
    yet_to_train_df = pd.DataFrame(yet_to_train) if yet_to_train else pd.DataFrame()

    # Get department data - training by department
    dept_data = conn.execute("""
        SELECT e.department,
               COUNT(DISTINCT e.id) as total_staff,
               COUNT(DISTINCT CASE WHEN t.id IS NOT NULL THEN e.id END) as trained_staff,
               COUNT(CASE WHEN strftime('%Y', t.training_date) = ? THEN t.id END) as total_trainings
        FROM employees e
        LEFT JOIN trainings t ON e.id = t.employee_id
                              AND t.deleted_at IS NULL
                              AND strftime('%Y', t.training_date) = ?
        WHERE e.deleted_at IS NULL
          AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
          AND e.department IS NOT NULL
          AND e.department != ''
        GROUP BY e.department
        HAVING total_trainings > 0
        ORDER BY total_trainings DESC
    """, (str(report_year), str(report_year))).fetchall()
    dept_df = pd.DataFrame(dept_data, columns=['Department', 'Total Staff', 'Trained Staff', 'Total Trainings']) if dept_data else pd.DataFrame()
    # Filter out rows with zero or negative values
    if not dept_df.empty:
        dept_df = dept_df[(dept_df['Total Staff'] > 0) & (dept_df['Total Trainings'] > 0)]
        dept_df['Coverage Rate'] = (dept_df['Trained Staff'] / dept_df['Total Staff'] * 100).round(1)

    return {
        'year': report_year,
        'summary': {
            'total_trainings': total_trainings,
            'total_staff': total_staff,  # Master + Hires + Leavers
            'active_staff': active_staff,  # Master + Hires - Leavers
            'staff_trained': staff_trained,  # Active staff trained
            'stores_trained': stores_trained,
            'pending_trainings': pending_count,
        },
        'training_frequency': freq_df,
        'stores_coverage': {
            'total': total_stores_count,
            'trained': stores_trained,
            'pending': stores_pending,
            'trained_pct': (stores_trained/total_stores_count*100 if total_stores_count else 0),
            'pending_pct': (stores_pending/total_stores_count*100 if total_stores_count else 0),
        },
        'staff_participation': {
            'eligible': eligible,
            'participated': participated,
            'not_participated': not_participated,
            'participation_rate': (participated/eligible*100 if eligible else 0),
        },
        'food_safety': {
            'required': food_report['required'],
            'tested': food_report['tested'],
            'never_tested': food_report['required'] - food_report['tested'],
            'tested_pct': (food_report['tested']/food_report['required']*100 if food_report['required'] else 0),
            'due_count': food_report['due_count'],
            'due_details': food_report['due_details'],
        },
        'compliance': compliance_data,
        'venue_analysis': venue_df,
        'regional_coverage': region_df,
        'department_coverage': dept_df,
        'monthly_trends': monthly_df,
        'staff_yet_to_train': yet_to_train_df,
    }

with col_dl1:
    if st.button("📊 Generate PowerPoint", use_container_width=True, type="primary"):
        with st.spinner("Generating PowerPoint presentation..."):
            report_data = prepare_report_data()
            pptx_bytes = create_pptx_report(report_data)
            st.download_button(
                "⬇️ Download PowerPoint",
                data=pptx_bytes,
                file_name=f"training_report_{report_year}.pptx",
                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                use_container_width=True
            )

with col_dl2:
    if st.button("📄 Generate Word Document", use_container_width=True, type="primary"):
        with st.spinner("Generating Word document..."):
            report_data = prepare_report_data()
            word_bytes = create_word_report(report_data)
            st.download_button(
                "⬇️ Download Word Document",
                data=word_bytes,
                file_name=f"training_report_{report_year}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                use_container_width=True
            )

with col_dl3:
    if st.button("📕 Generate PDF", use_container_width=True, type="primary"):
        with st.spinner("Generating PDF report..."):
            report_data = prepare_report_data()
            pdf_bytes = create_pdf_report(report_data)
            st.download_button(
                "⬇️ Download PDF",
                data=pdf_bytes,
                file_name=f"training_report_{report_year}.pdf",
                mime="application/pdf",
                use_container_width=True
            )

st.markdown("---")

# Create expandable sections for each report
with st.expander("🎯 Training Frequency Report", expanded=False):
    st.subheader(f"Training Sessions Held in {report_year}")
    freq_data = training_frequency_report(conn, year=report_year, filters=filters)

    if freq_data:
        freq_df = pd.DataFrame(freq_data, columns=[
            "Training Title", "Times Held", "All Dates", "First Date", "Last Date"
        ])

        st.dataframe(freq_df, use_container_width=True)

        # Chart
        fig_freq = px.bar(freq_df, x="Training Title", y="Times Held",
                         title=f"Training Frequency in {report_year}",
                         text="Times Held")
        fig_freq.update_layout(margin=dict(l=30,r=10,b=180,t=60), xaxis_tickangle=-45)
        st.plotly_chart(fig_freq, use_container_width=True)

        st.download_button(
            "Download Training Frequency Report (Excel)",
            data=to_excel_bytes({"Training Frequency": freq_df}),
            file_name=f"training_frequency_{report_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info(f"No training data found for {report_year}.")

with st.expander("🏪 Stores Training Coverage", expanded=False):
    st.subheader(f"Store Participation in {report_year}")

    stores_trained = stores_trained_count(conn, year=report_year, filters=filters)
    stores_pending = stores_pending_training(conn, year=report_year, filters=filters)
    total_stores_count = stores_trained + stores_pending

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Stores", total_stores_count)
    with col2:
        st.metric("Stores Trained", stores_trained,
                 delta=f"{(stores_trained/total_stores_count*100 if total_stores_count else 0):.1f}%")
    with col3:
        st.metric("Stores Pending", stores_pending,
                 delta=f"{(stores_pending/total_stores_count*100 if total_stores_count else 0):.1f}%",
                 delta_color="inverse")

    # Pie chart
    store_data = pd.DataFrame({
        "Status": ["Trained", "Pending"],
        "Count": [stores_trained, stores_pending]
    })
    fig_stores = px.pie(store_data, names="Status", values="Count",
                       title=f"Store Training Coverage {report_year}",
                       color_discrete_map={"Trained": "#00CC96", "Pending": "#EF553B"})
    st.plotly_chart(fig_stores, use_container_width=True)

with st.expander("👥 Staff Participation Analysis", expanded=False):
    st.subheader(f"Staff Training Participation in {report_year}")

    eligible = staff_eligible_for_training(conn, filters)
    participated = staff_participated_in_training(conn, year=report_year, filters=filters)
    not_participated = eligible - participated

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Eligible for Training", eligible)
    with col2:
        st.metric("Participated", participated,
                 delta=f"{(participated/eligible*100 if eligible else 0):.1f}%")
    with col3:
        st.metric("Not Participated", not_participated,
                 delta=f"{(not_participated/eligible*100 if eligible else 0):.1f}%",
                 delta_color="inverse")
    with col4:
        participation_rate = (participated/eligible*100 if eligible else 0)
        st.metric("Participation Rate", f"{participation_rate:.1f}%")

    # Bar chart
    participation_data = pd.DataFrame({
        "Category": ["Participated", "Not Participated"],
        "Count": [participated, not_participated]
    })
    fig_participation = px.bar(participation_data, x="Category", y="Count",
                              title=f"Staff Participation Overview {report_year}",
                              color="Category",
                              color_discrete_map={"Participated": "#00CC96", "Not Participated": "#EF553B"})
    st.plotly_chart(fig_participation, use_container_width=True)

with st.expander("⏰ Pending Trainings & Staff Yet to be Trained", expanded=False):
    st.subheader("Training Status and Pending Actions")

    pending_count = trainings_pending_count(conn, filters)
    yet_to_train = staff_yet_to_be_trained(conn, filters)

    col1, col2 = st.columns(2)
    with col1:
        st.metric("Trainings Pending (Due/Overdue)", pending_count)
    with col2:
        st.metric("Staff Yet to be Trained", len(yet_to_train))

    if yet_to_train:
        st.subheader("Staff Missing Required Trainings (Not on Probation)")
        yet_df = pd.DataFrame(yet_to_train)
        st.dataframe(yet_df, use_container_width=True, height=400)

        st.download_button(
            "Download Staff Yet to be Trained (Excel)",
            data=to_excel_bytes({"Staff Yet to Train": yet_df}),
            file_name="staff_yet_to_be_trained.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.success("✅ All eligible staff have received required trainings!")

with st.expander("🍽️ Food Handlers / Food Safety Detailed Report", expanded=True):
    st.subheader("Food Safety Training Compliance Report")

    food_report = food_handlers_report(conn, filters)

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Required to be Tested", food_report["required"])
    with col2:
        st.metric("Already Tested", food_report["tested"],
                 delta=f"{(food_report['tested']/food_report['required']*100 if food_report['required'] else 0):.1f}%")
    with col3:
        never_tested = food_report["required"] - food_report["tested"]
        st.metric("Never Tested", never_tested,
                 delta=f"{(never_tested/food_report['required']*100 if food_report['required'] else 0):.1f}%",
                 delta_color="inverse")
    with col4:
        st.metric("Due for Testing", food_report["due_count"])

    # Detailed table of staff due for testing
    if food_report["due_details"]:
        st.subheader("Staff Due for Food Safety Testing (Next 30 Days or Overdue)")
        due_df = pd.DataFrame(food_report["due_details"])

        # Color code the status column
        st.dataframe(due_df, use_container_width=True, height=400)

        # Summary by status
        status_summary = due_df.groupby("status").size().reset_index(name="Count")
        fig_food = px.bar(status_summary, x="status", y="Count",
                         title="Food Safety Testing Status Breakdown",
                         color="status")
        st.plotly_chart(fig_food, use_container_width=True)

        st.download_button(
            "Download Food Safety Report (Excel)",
            data=to_excel_bytes({
                "Food Safety Summary": pd.DataFrame([food_report]).drop(columns=["due_details"]),
                "Due for Testing": due_df
            }),
            file_name=f"food_safety_report_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.success("✅ All staff are up to date with Food Safety training!")

with st.expander("🎯 Compliance Dashboard - All 7 Required Trainings", expanded=False):
    st.subheader("Required Trainings Compliance Status")

    # Get compliance data for all 7 required trainings
    compliance_data = []

    for training_title in REQUIRED_TRAININGS:
        # Count employees with this training
        training_count = conn.execute("""
            SELECT COUNT(DISTINCT t.employee_id)
            FROM trainings t
            JOIN employees e ON e.id = t.employee_id
            WHERE t.deleted_at IS NULL
              AND e.deleted_at IS NULL
              AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
              AND lower(t.training_title) LIKE ?
        """, (f'%{training_title.split()[0].lower()}%',)).fetchone()[0]

        eligible = staff_eligible_for_training(conn, filters)
        compliance_rate = (training_count / eligible * 100) if eligible else 0

        # Count overdue/due
        overdue = conn.execute("""
            SELECT COUNT(DISTINCT t.employee_id)
            FROM trainings t
            JOIN employees e ON e.id = t.employee_id
            WHERE t.deleted_at IS NULL
              AND e.deleted_at IS NULL
              AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
              AND lower(t.training_title) LIKE ?
              AND date(t.next_due_date) < date('now')
        """, (f'%{training_title.split()[0].lower()}%',)).fetchone()[0]

        compliance_data.append({
            'Training': training_title,
            'Trained Staff': training_count,
            'Eligible Staff': eligible,
            'Compliance Rate': f"{compliance_rate:.1f}%",
            'Overdue': overdue
        })

    compliance_df = pd.DataFrame(compliance_data)
    st.dataframe(compliance_df, use_container_width=True, height=300)

    # Compliance chart
    compliance_df['Compliance_Numeric'] = compliance_df['Compliance Rate'].str.rstrip('%').astype(float)
    fig_compliance = px.bar(compliance_df, x='Training', y='Compliance_Numeric',
                           title="Compliance Rate by Training Type",
                           labels={'Compliance_Numeric': 'Compliance Rate (%)'},
                           color='Compliance_Numeric',
                           color_continuous_scale='RdYlGn',
                           range_color=[0, 100])
    fig_compliance.update_layout(margin=dict(l=30,r=10,b=200,t=60), xaxis_tickangle=-45)
    st.plotly_chart(fig_compliance, use_container_width=True)

    st.download_button(
        "Download Compliance Report (Excel)",
        data=to_excel_bytes({"Compliance Dashboard": compliance_df.drop(columns=['Compliance_Numeric'])}),
        file_name=f"compliance_dashboard_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with st.expander("📍 Training by Venue Analysis", expanded=False):
    st.subheader(f"Training Venues in {report_year}")

    venue_data = conn.execute("""
        SELECT t.training_venue,
               COUNT(*) as sessions,
               COUNT(DISTINCT t.employee_id) as unique_staff,
               COUNT(DISTINCT e.store) as unique_stores
        FROM trainings t
        JOIN employees e ON e.id = t.employee_id
        WHERE t.deleted_at IS NULL
          AND e.deleted_at IS NULL
          AND strftime('%Y', t.training_date) = ?
          AND t.training_venue IS NOT NULL
          AND t.training_venue != ''
        GROUP BY t.training_venue
        ORDER BY sessions DESC
    """, (str(report_year),)).fetchall()

    if venue_data:
        venue_df = pd.DataFrame(venue_data, columns=['Venue', 'Sessions Held', 'Unique Staff', 'Unique Stores'])
        st.dataframe(venue_df, use_container_width=True)

        # Venue chart
        fig_venue = px.bar(venue_df, x='Venue', y='Sessions Held',
                          title=f"Training Sessions by Venue ({report_year})",
                          text='Sessions Held')
        fig_venue.update_layout(margin=dict(l=30,r=10,b=180,t=60), xaxis_tickangle=-45)
        st.plotly_chart(fig_venue, use_container_width=True)

        st.download_button(
            "Download Venue Analysis (Excel)",
            data=to_excel_bytes({"Venue Analysis": venue_df}),
            file_name=f"venue_analysis_{report_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info(f"No venue data available for {report_year}.")

with st.expander("🌍 Regional Training Coverage", expanded=False):
    st.subheader("Training Coverage by Region")

    region_data = conn.execute("""
        SELECT COALESCE(t.region, e.region, 'Unknown') as region_name,
               COUNT(DISTINCT e.id) as total_staff,
               COUNT(DISTINCT CASE WHEN t.id IS NOT NULL THEN e.id END) as trained_staff,
               COUNT(CASE WHEN strftime('%Y', t.training_date) = ? THEN t.id END) as total_trainings
        FROM employees e
        LEFT JOIN trainings t ON e.id = t.employee_id
                              AND t.deleted_at IS NULL
                              AND strftime('%Y', t.training_date) = ?
        WHERE e.deleted_at IS NULL
          AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
        GROUP BY region_name
        HAVING total_trainings > 0
        ORDER BY total_trainings DESC
    """, (str(report_year), str(report_year))).fetchall()

    if region_data:
        region_df = pd.DataFrame(region_data, columns=['Region', 'Total Staff', 'Trained Staff', 'Total Trainings'])
        # Filter out rows with zero or negative values
        region_df = region_df[(region_df['Total Staff'] > 0) & (region_df['Total Trainings'] > 0)]
        region_df['Coverage Rate'] = (region_df['Trained Staff'] / region_df['Total Staff'] * 100).round(1)

        st.dataframe(region_df, use_container_width=True)

        # Regional coverage chart
        fig_regional = px.bar(region_df, x='Region', y=['Trained Staff', 'Total Staff'],
                             title="Regional Training Coverage",
                             barmode='group')
        fig_regional.update_layout(margin=dict(l=30,r=10,b=120,t=60), xaxis_tickangle=-45)
        st.plotly_chart(fig_regional, use_container_width=True)

        # Training by Region bar chart
        fig_regional_trainings = px.bar(region_df.sort_values('Total Trainings', ascending=False),
                                       x='Region', y='Total Trainings',
                                       title="Total Trainings by Region",
                                       text='Total Trainings')
        fig_regional_trainings.update_layout(margin=dict(l=30,r=10,b=120,t=60), xaxis_tickangle=-45)
        st.plotly_chart(fig_regional_trainings, use_container_width=True)

        st.download_button(
            "Download Regional Coverage (Excel)",
            data=to_excel_bytes({"Regional Coverage": region_df}),
            file_name="regional_coverage.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info("No regional data available.")

with st.expander("💼 Position-Based Training Analysis", expanded=False):
    st.subheader("Training by Employee Position")

    position_data = conn.execute("""
        SELECT e.position,
               COUNT(DISTINCT e.id) as total_staff,
               COUNT(DISTINCT CASE WHEN t.id IS NOT NULL THEN e.id END) as trained_staff,
               COUNT(DISTINCT t.id) as total_trainings
        FROM employees e
        LEFT JOIN trainings t ON e.id = t.employee_id AND t.deleted_at IS NULL
        WHERE e.deleted_at IS NULL
          AND (e.end_date IS NULL OR date(e.end_date) > date('now'))
          AND e.position IS NOT NULL
          AND e.position != ''
        GROUP BY e.position
        ORDER BY total_staff DESC
        LIMIT 20
    """).fetchall()

    if position_data:
        position_df = pd.DataFrame(position_data, columns=['Position', 'Total Staff', 'Trained Staff', 'Total Trainings'])
        position_df['Training per Staff'] = (position_df['Total Trainings'] / position_df['Total Staff']).round(2)

        st.dataframe(position_df, use_container_width=True)

        # Position chart
        fig_position = px.bar(position_df.head(15), x='Position', y='Training per Staff',
                             title="Top 15 Positions by Training per Staff",
                             text='Training per Staff')
        fig_position.update_layout(margin=dict(l=30,r=10,b=200,t=60), xaxis_tickangle=-65)
        st.plotly_chart(fig_position, use_container_width=True)

        st.download_button(
            "Download Position Analysis (Excel)",
            data=to_excel_bytes({"Position Analysis": position_df}),
            file_name="position_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info("No position data available.")

with st.expander("📅 Monthly Training Trends", expanded=False):
    st.subheader(f"Monthly Training Activity in {report_year}")

    monthly_data = conn.execute("""
        SELECT strftime('%m', t.training_date) as month,
               COUNT(*) as trainings,
               COUNT(DISTINCT t.employee_id) as unique_staff,
               COUNT(DISTINCT e.store) as unique_stores
        FROM trainings t
        JOIN employees e ON e.id = t.employee_id
        WHERE t.deleted_at IS NULL
          AND e.deleted_at IS NULL
          AND strftime('%Y', t.training_date) = ?
        GROUP BY month
        ORDER BY month
    """, (str(report_year),)).fetchall()

    if monthly_data:
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        monthly_df = pd.DataFrame(monthly_data, columns=['Month', 'Trainings', 'Unique Staff', 'Unique Stores'])
        monthly_df['Month Name'] = monthly_df['Month'].apply(lambda x: month_names[int(x)-1])

        st.dataframe(monthly_df[['Month Name', 'Trainings', 'Unique Staff', 'Unique Stores']], use_container_width=True)

        # Monthly trends chart
        fig_monthly = px.line(monthly_df, x='Month Name', y=['Trainings', 'Unique Staff'],
                             title=f"Monthly Training Trends ({report_year})",
                             markers=True)
        fig_monthly.update_layout(margin=dict(l=30,r=10,b=80,t=60))
        st.plotly_chart(fig_monthly, use_container_width=True)

        st.download_button(
            "Download Monthly Trends (Excel)",
            data=to_excel_bytes({"Monthly Trends": monthly_df[['Month Name', 'Trainings', 'Unique Staff', 'Unique Stores']]}),
            file_name=f"monthly_trends_{report_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info(f"No monthly data available for {report_year}.")

with st.expander("🏢 Department-Wise Pending Trainings", expanded=False):
    st.subheader("Pending Trainings by Department")

    # Get flags and organize by department
    flags = employees_due_flags(conn, filters)

    dept_pending = {}
    for emp_id, status in flags.items():
        if status.get("any_due_30", False) or status.get("any_overdue", False):
            # Get employee department
            emp = conn.execute("SELECT department FROM employees WHERE id=?", (emp_id,)).fetchone()
            if emp and emp[0]:
                dept = emp[0]
                if dept not in dept_pending:
                    dept_pending[dept] = {'due_30': 0, 'overdue': 0, 'missing': 0}

                if status.get("any_overdue", False):
                    dept_pending[dept]['overdue'] += 1
                elif status.get("any_due_30", False):
                    dept_pending[dept]['due_30'] += 1

                if status.get("any_missing", False):
                    dept_pending[dept]['missing'] += 1

    if dept_pending:
        pending_df = pd.DataFrame([
            {'Department': dept, 'Due in 30 Days': data['due_30'], 'Overdue': data['overdue'], 'Missing Trainings': data['missing']}
            for dept, data in dept_pending.items()
        ]).sort_values('Overdue', ascending=False)

        st.dataframe(pending_df, use_container_width=True)

        # Pending trainings chart
        fig_pending = px.bar(pending_df, x='Department', y=['Due in 30 Days', 'Overdue', 'Missing Trainings'],
                            title="Pending Trainings by Department",
                            barmode='stack')
        fig_pending.update_layout(margin=dict(l=30,r=10,b=160,t=60), xaxis_tickangle=-45)
        st.plotly_chart(fig_pending, use_container_width=True)

        st.download_button(
            "Download Department Pending Report (Excel)",
            data=to_excel_bytes({"Department Pending": pending_df}),
            file_name="department_pending_trainings.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.success("✅ No pending trainings across all departments!")

with st.expander("👋 Trained Employees Who Left", expanded=False):
    st.subheader("Training ROI Analysis - Employees Who Left After Training")
    st.caption("Shows employees who received training but have since left the company. Useful for understanding training ROI and retention patterns.")

    left_data = trained_employees_who_left(conn, filters)

    if left_data:
        left_df = pd.DataFrame(left_data, columns=[
            'ID', 'Name', 'Email', 'Department', 'Store', 'Position', 'Region',
            'Start Date', 'End Date', 'Total Trainings', 'Training Titles',
            'First Training', 'Last Training'
        ])

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Employees Who Left After Training", len(left_df))
        with col2:
            total_trainings_lost = left_df['Total Trainings'].sum()
            st.metric("Total Trainings Lost", int(total_trainings_lost))
        with col3:
            avg_trainings = left_df['Total Trainings'].mean()
            st.metric("Avg Trainings per Leaver", f"{avg_trainings:.1f}")
        with col4:
            # Calculate retention period (days between first training and end date)
            left_df['Start Date'] = pd.to_datetime(left_df['Start Date'], errors='coerce')
            left_df['End Date'] = pd.to_datetime(left_df['End Date'], errors='coerce')
            left_df['First Training'] = pd.to_datetime(left_df['First Training'], errors='coerce')
            left_df['Retention Days'] = (left_df['End Date'] - left_df['First Training']).dt.days
            avg_retention = left_df['Retention Days'].mean()
            st.metric("Avg Retention After 1st Training", f"{avg_retention:.0f} days" if pd.notna(avg_retention) else "N/A")

        st.dataframe(left_df, use_container_width=True, height=400)

        # Breakdown by department
        dept_breakdown = left_df.groupby('Department').agg({
            'ID': 'count',
            'Total Trainings': 'sum'
        }).reset_index()
        dept_breakdown.columns = ['Department', 'Employees Who Left', 'Total Trainings Lost']
        dept_breakdown = dept_breakdown.sort_values('Employees Who Left', ascending=False)

        st.subheader("Breakdown by Department")
        fig_dept_left = px.bar(dept_breakdown, x='Department', y=['Employees Who Left', 'Total Trainings Lost'],
                              title="Trained Employees Who Left - By Department",
                              barmode='group')
        fig_dept_left.update_layout(margin=dict(l=30,r=10,b=160,t=60), xaxis_tickangle=-45)
        st.plotly_chart(fig_dept_left, use_container_width=True)

        st.download_button(
            "Download Trained Employees Who Left (Excel)",
            data=to_excel_bytes({
                "Employees Who Left": left_df,
                "Department Breakdown": dept_breakdown
            }),
            file_name="trained_employees_who_left.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.success("✅ No trained employees have left the company!")

st.header("📈 Interactive Analytics")

# Pull filtered datasets once
df_tr = get_trainings_df(conn, filters)
df_emps = get_employees_df(conn, filters)

tabs = st.tabs([
    "Dept", "Title", "Trend", "Coverage", "Due/Overdue",
    "Region", "Stores", "Heatmap", "Stacked Titles", "Joins vs Leaves", "Treemap"
])

with tabs[0]:
    fig, data = chart_trainings_by_department(df_tr)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=480)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Trainings by Department": data}),
                           file_name="trainings_by_department.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[1]:
    fig, data = chart_trainings_by_title(df_tr)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=480)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Trainings by Title": data}),
                           file_name="trainings_by_title.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[2]:
    fig, data = chart_trainings_by_month(df_tr)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=420)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Trainings by Month": data}),
                           file_name="trainings_by_month.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[3]:
    fig, data = chart_coverage_by_department(conn, df_tr, filters)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=480)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Coverage by Department": data}),
                           file_name="coverage_by_department.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[4]:
    fig, data = chart_due_overdue_by_dept(conn, filters)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=480)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Due vs Overdue by Department": data}),
                           file_name="due_overdue_by_department.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[5]:
    fig, data = chart_region_pie(df_tr)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=480)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Trainings by Region": data}),
                           file_name="trainings_by_region.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[6]:
    topn = st.slider("Top stores", min_value=5, max_value=50, value=20, step=1)
    fig, data = chart_top_stores(df_tr, topn=topn)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=520)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Top Stores": data}),
                           file_name="top_stores.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[7]:
    fig, data = chart_heat_dept_month(df_tr)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=520)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Dept x Month Heatmap": data}),
                           file_name="dept_month_heatmap.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[8]:
    k = st.slider("Top training titles", min_value=3, max_value=10, value=6, step=1)
    fig, data = chart_stacked_titles_by_dept(df_tr, top_titles=k)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=520)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Stacked Titles by Dept": data}),
                           file_name="stacked_titles_by_dept.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[9]:
    dfrom = filters.get("date_from") or f"{date.today().year}-01-01"
    dto   = filters.get("date_to") or date.today().isoformat()
    fig, data = chart_joins_vs_leaves_by_month(conn, filters, date_from=dfrom, date_to=dto)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=440)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Joins vs Leaves by Month": data}),
                           file_name="joins_vs_leaves_by_month.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")

with tabs[10]:
    fig, data = chart_treemap_store_coverage(conn, df_tr, filters)
    if fig:
        st.plotly_chart(fig, use_container_width=True, height=520)
        st.download_button("Download data (Excel)", data=to_excel_bytes({"Store Coverage": data}),
                           file_name="store_coverage.xlsx", use_container_width=True)
    else:
        st.info("No data to display.")
